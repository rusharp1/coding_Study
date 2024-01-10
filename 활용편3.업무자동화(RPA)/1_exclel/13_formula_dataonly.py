from openpyxl import load_workbook
wb = load_workbook("sample_formula.xlsx")
ws = wb.active

# 수석 출력
for row in ws.values:
    for cell in row:
        print(cell)


wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

# 수식이 아닌, 계산된 값 출력.
# evalluate 되지 않은 데이터는 None이라고 표시.
for row in ws.values:
    for cell in row:
        print(cell)