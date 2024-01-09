# 파일 불러오기 라이브러리 import
from openpyxl import load_workbook

# sample.xlsx 파일에서 wb를 불러옴
wb = load_workbook("sample.xlsx") 
ws = wb.active # 활성화된 sheet

for x in range(1,11):
    for y in range(1, 11):
        # ws.cell(row=x, column=y, value=randint(0,100))
        print(ws.cell(row=x, column=y).value, end = ", ")
    print()

# cell 개수를 모를때
for x in range (1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end = ", ")
    print()