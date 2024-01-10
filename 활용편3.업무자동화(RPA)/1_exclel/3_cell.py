from openpyxl import Workbook
from random import *

wb=Workbook()
# wb.active
# 새로운 Sheet기본 이름으로 생성.
ws = wb.active
ws.title = "Mysheet"


colm = "A"
for i in range(2):
    colm = chr(ord(colm) + i)
    for n in range(1, 11):
        ws["{}{}".format(colm,n)] = i*10 + n


### 특정 셀의 값 or 정보 출력
        
# A1의 셀 정보를 출력.
print(ws["A1"])
# A1의 값 정보를 출력.
print(ws["A1"].value)
print(ws["C1"].value)

# row = 1,2,3 ...
# column = A, B, C ...
# A1의 값
print("A1 =" ,ws.cell(row=1, column=1).value)
# B1의 값
print("B1 - ",ws.cell(row=1, column=2).value)

### 특정 셀에 값 입력하기.
ws.cell(column=3, row=1, value=30)
print("C1.value = ", ws["C1"].value)
ws["C1"].value=29
print("C1.value = ", ws["C1"].value)
ws["C1"] = 28
print("C1.value = ", ws["C1"].value)

# 반복문을 이용해 랜덤 숫자 채우기
for x in range(1,11):
    for y in range(1, 11):
        ws.cell(row=x, column=y, value=randint(0,100))
        # ws.cell(row=x, column=y, value = (x-1)*10 + y)

wb.save("sample.xlsx")