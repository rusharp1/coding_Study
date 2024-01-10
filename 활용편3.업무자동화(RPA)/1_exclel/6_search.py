from openpyxl import load_workbook

# 기존 Excel 파일 불러오기
wb = load_workbook(filename='sample.xlsx')

# 활성화된 시트 가져오기
ws = wb.active

# 데이터 읽기
for row in ws.iter_rows(min_row = 2):
    # 번호, 영어, 수학
    if int(row[1].value) > 80:
        print("{}번 학생은 영어 점수가 80점을 초과합니다.".format(row[0].value))

for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "영어":
            cell.value = "컴퓨터"

for cell in ws[ws.max_row]:
    cell.value = ""

# 워크북 저장
wb.save('sample.xlsx')