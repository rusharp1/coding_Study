from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 병합하기
# B2~D2까지 합치기
ws.merge_cells("B2:D2")
ws["B2"].value = "Merged Cell"

ws.merge_cells("B3:D3")
ws["B3"].value = "Merged Cell2"


ws.unmerge_cells("B3:D3")
wb.save("sample_merge.xlsx")
