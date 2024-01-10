import datetime
from openpyxl import  Workbook

wb = Workbook()
ws = wb.active

# 함수명을 넣어 excel 수식을 설정할 수 있음.
ws["A1"] = datetime.datetime.today()
ws["A2"] = "=SUM(1,2,3)"
ws["A3"] = "=AVERAGE(1,2,3)"

ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "=SUM(A4:A5)"

wb.save("sample_formula.xlsx")