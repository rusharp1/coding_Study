from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_from_string
from random import *
wb = Workbook()
ws = wb.active

# 1줄씩 data 넣기
ws.append(["번호","영어","수학"])
for i in range(1,11) : # 10개 데이터 넣기
    ws.append([i, randint(0,100), randint(0,100)])

# B컬럼만 가져오기.
col_B = ws["B"]
col_B = [b.value for b in col_B] 

# B, C 컬럼 가져오기
col_range = ws["B:C"]
col_range = [cell.value for cols in col_range for cell in cols ]
# print(col_range)
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

# 첫번째 row만 가져오기
row_title = ws[1]
# for cell in row_title:
#     print(cell.value)

# 2번째 행부터 6번재 행까지 가져오기
row_range = ws[2:6]
row_range = [cell.value for rows in row_range for cell in rows ]
# print(row_range)

# 2번재 행부터 마지막 행까지
row_range = ws[2:ws.max_row]
for rows in row_range:
    for cell in rows:
        # cell의 value 출력 # 1, n, m ...
        # print(cell.value, end = " " )
        # cell의 좌표값 출력 A2, B2, C2 ...
        # print(cell.coordinate, end = " ")
        # cell의 좌표값을 tuple값으로 출력 ('A', 2) ('B', 2) ('C', 2) ...
        xy = coordinate_from_string(cell.coordinate)
        # xy[0] = 'A', xy[1] = 2 ... 
        print(xy, end = " ")
    print()

wb.save("sample.xlsx")