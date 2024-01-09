from openpyxl import Workbook
wb=Workbook()
# wb.active
# 새로운 Sheet기본 이름으로 생성.
ws = wb.create_sheet()
# 시트 이름 변경.
ws.title = "Mysheet"
# RGB형태로 값을 넣어주면 탭 색상 변경.
ws.sheet_properties.tabColor = "ffcccc"

# 시트 맨마지막에 주어진 이름으로 시트 생성.
ws1 = wb.create_sheet("YourSheet")
# 시트 위치를 지정하여 시트 생성.
ws2 = wb.create_sheet("Third Sheet", 2)

# 시트명을 통해 특정 시트에 접근.
new_ws = wb["Third Sheet"]
# 시트명을 모두 출력함.
print(wb.sheetnames)


# 시트 복사
# A1 위치에 "test" 입력
new_ws["A1"] = "test"
target = wb.copy_worksheet(new_ws)
# 복사된 시트 이름을 지정함.
# 지정하지 않는경우, new_ws.title + Copied 로 이름지정됨
target.title = "Copied Sheet"

wb.save("sample.xlsx")