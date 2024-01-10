from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호, 영어, 수학
a1 = ws['A1']
b1 = ws['B1']
c1 = ws['C1']
d1 = ws['D1']
e1 = ws['E1']

# A열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5
# 1행의 높이를 20로 설정
ws.row_dimensions[1].height = 50

# 글자 빨간색, 이탤릭, 두껍게
a1.font = Font(color="FF0000", italic=True, bold=True)
# 폰트명 : Arial, 취소선
b1.font = Font(color="cc33ff", name="Arial", strike=True)
# 글자 크기 20, 밑줄 적용.
c1.font = Font(color="0000ff", size=20, underline="single")


# 테두리 적용
thin_border = Border(left = Side(style="thin"),
                     right = Side(style="thin"),
                     top = Side(style="thin"),
                     bottom = Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90점 넘는 셀에 대해 초록색으로 적용.
for row in ws.rows:
    for cell in row:
        # 모든셀을 중앙정렬로 변경.
        # center, bottom, top, left, rignt 등이 있음.
        cell.alignment = Alignment(horizontal = "center", 
                                   vertical = "center")
        # 번호일 때에는  넘어감
        if cell.column == 1:
            continue
        # 숫자일 때만 적용됨.
        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor= "22ff22", fill_type="solid")
            cell.font = Font(color="FF0000")
# B2기준으로 틀 고정.
# B2 기준 왼쪽, 오른쪽으로 틀 생성.
ws.freeze_panes = "B2"
wb.save("sample.xlsx")
