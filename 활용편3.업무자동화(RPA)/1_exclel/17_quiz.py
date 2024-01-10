# 출석     : 10
# 퀴즈1    : 10
# 퀴즈2    : 10
# 중간고사 : 20
# 기말고사 : 30
# 프로젝트 : 20
# 총합     : 100

# 마지막수업을 모두 마치고 이번 학기 학생들의최종 성적을 검토하는 과정에서
# 퀴즈2문제에 오류를 발견하여 모두 만점 처리를 하기로 하였습니다.
# 현재까지 작성된 최종 성적 데이터를기준으로 아래와 같이 수정하시오.

# 1. 퀴즈2 점수를 10으로 수정
# 2. H열에 총점(SUM 이용), I열에 성적 정보 추가
# - 총점 90점 이상은 A, 80점 이상은 B, 70점 이상은 C, 나머지는 D
# 3. 출석이 5 미만인 학생은 총점 관계없이 False

# ※ 최종 파일명 : score.xlsx

from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

ws.append(["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"])
for i in range(1, 11):
    ws.append([i, randint(1,10), randint(1,10), randint(1,10),
               randint(1,20), randint(1,30), randint(1,20)])

for cell in ws[1]:
    if cell.value == "퀴즈2":
        print(cell)
        column = cell.column
        break

for i in range(2, ws.max_row+1):
    ws.cell(column = column,row = i, value = 10)

# H열에 총점, I열에 성적 정보 추가하기
ws["H1"] = "총점"
ws["I1"] = "성적"

for i in range(2, ws.max_row +1):
    ws.cell(row = i, column = 8, value = "=SUM(B{}:G{})".format(i, i))
    ws.cell(row = i, column = 9,
            value = f'=IF(B{i} < 5, "F", IF(H{i} >= 90, "A", IF(H{i} >= 80, "B", IF(H{i} >= 70, "C", "D"))))')

wb.save("score.xlsx")