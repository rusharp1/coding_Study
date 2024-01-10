from openpyxl import load_workbook
from random import *

wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호, 컴퓨터, 수학
# 번호, 국어, 영어, 컴퓨터,  수학 으로 변경.

# B1~C11을 복사하여 행은 그대로, 열은 우측으로 2방향 이동함.
ws.move_range("B1:C11", rows = 0 ,cols = 2)
ws["B1"].value = "국어"
ws["C1"].value = "영어"

for rows in range(2,11) : # 10개 데이터 넣기
    for cols in range(2,4):
        ws.cell(row=rows, column=cols, value=randint(0,100))

wb.save("sample.xlsx")